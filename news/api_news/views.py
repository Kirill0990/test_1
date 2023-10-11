from rest_framework import viewsets
from datetime import datetime
from openpyxl import Workbook
from django.http import HttpResponse
import csv

from .serializers import PostSerializer
from posts.models import Post


class PostViewSet(viewsets.ModelViewSet):
    queryset = Post.objects.all()
    serializer_class = PostSerializer


def export_to_xlsx(request):

    queryset = Post.objects.all()

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename={date}-posts.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()

    worksheet = workbook.active
    worksheet.title = 'Post'

    columns = [
        'id',
        'title',
        'description',
        'pub_date',
    ]
    row_num = 1

    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    for post in queryset:
        row_num += 1

        row = [
            post.pk,
            post.title,
            post.description,
            post.pub_date,

        ]

        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)

    return response


def export_csv(request):
    # Your data retrieval logic goes here
    data = Post.objects.all()

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="post.csv"'

    writer = csv.writer(response)
    writer.writerow([
        'id',
        'title',
        'description',
        'pub_date',
        ])

    for post in data:
        writer.writerow([post.pk,
                         post.title,
                         post.description,
                         post.pub_date])

    return response
